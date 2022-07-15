# imports required library
import os
import openpyxl
import jdatetime
from pyromod import listen
from pyrogram import Client, filters
from pyrogram.types import ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton

# connection to telegram bot
api_id = 12721742  # your api id
api_hash = "2a81674bd5e1ccbaed8c07f898d614ca"  # your api hash
bot_token = "2018754108:AAHPAlIOiAQDh-mC8CyYMmB2An-oyQz5sy8"  # token your bot
app = Client("bot", api_id=api_id, api_hash=api_hash, bot_token=bot_token)

# variable required for begin project
file_information = "Information.xlsx"
file_saved_betters = "SavedAndBetter.xlsx"
file_all_user = "allUser.txt"
file_user_login = "userLogin.txt"
admin = 760148720  # user id admin
ABOUT_BOT = """این ربات در جهت پیشرفت و آبادانی روستایمان « کهن دیار ازغند » طراحی شده است. 
🤔 چنانچه در استفاده از ربات به مشکل برخورد کردید می‌توانید با مدیریت ربات یا طراح و توسعه دهنده ربات در ارتباط باشید💐
👌🏻 همچنین جهت سفارش و طراحی ربات خود می‌توانید با توسعه دهنده در ارتباط باشید
ایدی دهیار محترم: 
@Dehyar_SAFARI
🛠 توسعه دهنده : رضا بخش زایی
🆔 @REZABZ2 
📞 09154172849"""
LIST_WORDS = ["📝ثبت نام و ورود", "🤖درباره ربات", "📝ثبت پیشنهادات و انتقادات", "🖋ثبت طرح📝",
              "/start", "/Start", "/START",
              "👤نمایش کاربران👤", "📲ارسال پیام همگانی📲",
              "📩پیام های ذخیره شده📩", "🥇پیام های برتر🥇","✅بازگشت به منو اولیه♻️",
              "📄گزارشکار پروژه های در حال اجرا ⛓","🦹‍♂️ادامه در حالت ناشناس"]
question_one = """1️⃣ اگر شما دهیار روستا بودید چه اقداماتی را در راستای پیشرفت و آبادانی روستا انجام می‌دادید؟"""
question_two = """2️⃣ لطفا بزرگترین پتانسیل روستا را در چند خط توضیح بدهید."""
question_three = """3️⃣ انتظارات خود از دهیاری را بیان کنید."""
# keyboards
key_admin = ReplyKeyboardMarkup(
    [
        ["👤نمایش کاربران👤", "📲ارسال پیام همگانی📲"],
        ["📩پیام های ذخیره شده📩", "🥇پیام های برتر🥇"]
    ], resize_keyboard=True
)
key_user_not_login = ReplyKeyboardMarkup(
    [
        ["📝ثبت نام و ورود", "🦹‍♂️ادامه در حالت ناشناس"],
        ["📄گزارشکار پروژه های در حال اجرا ⛓"],
        ["🤖درباره ربات"]
    ], resize_keyboard=True
)
key_user_login = ReplyKeyboardMarkup(
    [
        ["📝ثبت پیشنهادات و انتقادات", "🖋ثبت طرح📝"],
        ["✅بازگشت به منو اولیه♻️"]
    ], resize_keyboard=True
)


# functions required
def required():
    try:
        workbook = openpyxl.load_workbook(file_information)
        workbook.close()
    except:
        workbook = openpyxl.Workbook()
        workbook.save(file_information)
    try:
        workbook = openpyxl.load_workbook(file_saved_betters)
        workbook.close()
    except:
        workbook = openpyxl.Workbook()
        workbook.save(file_saved_betters)
    try:
        file = open(file_all_user, "r", encoding="UTF-8").read()
    except:
        file_opened = open(file_all_user, "a", encoding="UTF-8")
        file_opened.close()
    try:
        file = open(file_user_login, "r", encoding="UTF-8").read()
    except:
        file_opened = open(file_user_login, "a", encoding="UTF-8")
        file_opened.close()


def check_exist_id(file_name, number_id):
    ids = open(file_name, "r", encoding="UTF-8").read().split()
    for one_id in ids:
        if int(number_id) == int(one_id):
            return 1
    return 0


def get_information(number_id, file_name):
    workbook = openpyxl.load_workbook(file_name, read_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if int(row[0]) != int(number_id):
            continue
        return row
    return 0


def save_information(file_name, information):
    last = 1
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet = workbook.active
    for i in sheet.iter_rows(max_row=0):
        last += 1
    sheet[f"A{last}"] = information[0]
    sheet[f"B{last}"] = information[1]
    sheet[f"C{last}"] = information[2]
    os.remove(file_name)
    workbook.save(filename=file_name)


def save_id(file_name, number_id):
    file = open(file_name, "a", encoding="UTF-8")
    file.write(str(number_id )+ " ")
    file.close()


def get_messages(number_id, data):
    workbook = openpyxl.load_workbook(file_saved_betters, read_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if int(number_id) == int(row[0]) and row[1] == data:
            return row
    return 0


# panel admin
@app.on_message(filters.user(admin) & filters.command("start", "/"))
async def start_admin(client, message):
    await client.send_message(admin, "💡سلام ، به پنل مدیریت خوش آمدید 👨🏻‍💻 ", reply_markup=key_admin)


@app.on_message(filters.user(admin) & filters.regex("^📲ارسال پیام همگانی📲$"))
async def message_to_all_user(client, message):
    ids = open(file_all_user, "r", encoding="UTF-8").read().split()
    if len(ids) < 1:
        await message.reply("📂")
        await message.reply("💢کاربری وجود ندارد .")
    else:
        await message.reply("📲به بخش ارسال پیام همگانی خوش امدید.", reply_markup=ReplyKeyboardMarkup(
            [
                ["بازگشت"]
            ], resize_keyboard=True
        ))
        mes = await client.ask(admin, "🔹پیام مورد نظر را وارد کنید.")
        if mes.text == "بازگشت":
            await mes.reply("✅به پنل مدیریت بازگشتید.♻️", reply_markup=key_admin)
        elif mes.text in LIST_WORDS:
            await  mes.reply("❌پیام وارد شده اشتباه است\n"
                             "لطفا دوباره سعی کنید و دقت کنید که پیام وارد شده به صورت متن باشد.",
                             reply_markup=key_admin)
        else:
            for one_id in ids:
                await client.send_message(int(one_id), f"**📍 از طرف دهیاری 📍**\n📝متن پیام: {mes.text}")
            await mes.reply("✅عملیات مورد نظر با موفقیت انجام شد.", reply_markup=key_admin)


@app.on_message(filters.user(admin) & filters.regex("^👤نمایش کاربران👤$"))
async def show_users(client, message):
    workbook = openpyxl.load_workbook(file_information, read_only=True)
    sheet = workbook.active
    text = ""
    switch = 1
    for row in sheet.iter_rows(values_only=True):
        text += f"📝نام و نام خانوادگی: {row[1]}\n📞شماره تلفن: {row[2]}\n➖➖➖➖➖➖➖➖➖➖➖➖\n"
        if 1000 < len(text) < 1250:
            await message.reply(f"**🧑🏼‍💻کاربران 🧑🏼‍💻**\n{text}")
            text = ""
            switch = 0
    if len(text) > 0:
        await message.reply(f"**🧑🏼‍💻کاربران 🧑🏼‍💻**\n{text}")
    elif switch == 1:
        await message.reply("📂")
        await message.reply("💢کاربری وجود ندارد .")
    workbook.close()


@app.on_message(filters.user(admin) & filters.regex("^📩پیام های ذخیره شده📩$"))
async def show_message_saved(client, message):
    workbook = openpyxl.load_workbook(file_saved_betters, read_only=True)
    sheet = workbook.active
    text = ""
    switch = 1
    for row in sheet.iter_rows(values_only=True):
        if row[1] == "save":
            text += f"**📝 متن پیام:** {row[2]}\n➖➖➖➖➖➖➖➖➖➖➖\n"
            if 1000 < len(text) < 1550:
                await message.reply(f"**📩پیام های ذخیره شده📩**\n{text}")
                text = ""
                switch = 0
    if len(text) > 0:
        await message.reply(f"**📩پیام های ذخیره شده📩**\n{text}")
    elif switch == 1:
        await message.reply("📂")
        await message.reply("💢پیامی موجود نیست.")
    workbook.close()


@app.on_message(filters.user(admin) & filters.regex("^🥇پیام های برتر🥇$"))
async def show_message_better(client, message):
    workbook = openpyxl.load_workbook(file_saved_betters, read_only=True)
    sheet = workbook.active
    text = ""
    switch = 1
    for row in sheet.iter_rows(values_only=True):
        if row[1] == "better":
            text += f"**📝 متن پیام:** {row[2]}\n➖➖➖➖➖➖➖➖➖➖➖\n"
            if 1000 < len(text) < 1550:
                await message.reply(f"**🥇پیام های برتر🥇**\n{text}")
                text = ""
                switch = 0
    if len(text) > 0:
        await message.reply(f"**🥇پیام های برتر🥇**\n{text}")
    elif switch == 1:
        await message.reply("📂")
        await message.reply("💢پیامی موجود نیست.")
    workbook.close()


@app.on_callback_query(filters.user(admin))
async def callback_query_admin(client, callback):
    data = callback.data
    callback_id = callback.message.id
    text = callback.message.text
    information = get_messages(callback_id, data)
    if information == 0:
        save_information(file_saved_betters, [callback_id, data, text])
        await callback.answer(f"🟢✔️این پیام در دسته  {data} قرار گرفت\n متن:\n{text}", show_alert=True)
    elif int(information[0]) == int(callback_id) and information[1] == data:
        await callback.answer(f"♦️این پیام در دسته  {data} قبلا قرار گرفته است.", show_alert=False)
    elif int(information[0]) == int(callback_id) and information[1] != data:
        save_information(file_saved_betters, [callback_id, data, text])
        await callback.answer(f"🟢✔️این پیام در دسته  {data} قرار گرفت\n متن:\n{text}", show_alert=True)


# panel users
@app.on_message(filters.private & filters.command("start", "/"))
async def start_user(client, message):
    await message.reply("""🤖سلام خوش آمدید
این ربات در جهت پیشرفت و آبادانی « کهن دیار ازغند » طراحی شده است. 
☑️ امید داریم با ارتباط بیشتر از این مسیر وارسال نظرات و انتقادات و پیشنهادات
 شما به پیشرفت هر چه بهتر دیارمان فزونی بخشیم🌺🙏""")
    if check_exist_id(file_all_user, message.from_user.id) == 0:
        save_id(file_all_user, message.from_user.id)
    if check_exist_id(file_user_login, message.from_user.id) == 0:
        await message.reply("✅📋لطفا اول به سه سوال زیر پاسخ دهید.")
        answer_one = await client.ask(message.from_user.id, question_one)
        answer_two = await client.ask(answer_one.from_user.id, question_two)
        answer_three = await client.ask(answer_two.from_user.id, question_three)
        await client.send_message(admin, f"**🗒پاسخ کاربر جدید به 3 سوال🗒**\n"
                                         f"{question_one}\n{answer_one.text}\n➖➖➖➖➖➖➖➖➖➖➖\n"
                                         f"{question_two}\n{answer_two.text}\n➖➖➖➖➖➖➖➖➖➖➖\n"
                                         f"{question_three}\n{answer_three.text}\n➖➖➖➖➖➖➖➖➖➖➖\n")
        await message.reply("""✅ پاسخ های شما ثبت شد.
💥شما به منو اصلی هدایت میشوید
همچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی
 بهترین نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""", reply_markup=key_user_not_login)
    else:
        information = get_information(message.from_user.id, file_information)
        await message.reply(f"""❇️سلام  {information[1]} عزیز\n❇️🖋شما قبلا به سه سوال پاسخ داده اید.
💥به منو اصلی هدایت میشوید
مچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی بهترین
 نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""",
                            reply_markup=key_user_login)


@app.on_message(filters.private & filters.regex("^🦹‍♂️ادامه در حالت ناشناس$"))
async def continuation_unknown(client, message):
    if check_exist_id(file_user_login, message.from_user.id) == 0:
        await message.reply("""🤖به بخش ناشناس خوش امدید.
    💥 بدون هیچ هویت انتقادات و پیشنهادات و طرح های مدنظر خود را با ما در میان بگذارید🙌""",
                            reply_markup=key_user_login)
    else:
        await message.reply("‼️کاربر محترم شما قبلا ثبت نام کردید و برای شما امکان ارسال پیام ناشناس نمیباشد.‼️\n"
                            "💢به صورت شناخته شده طرح یا انتقاد و پیشنهاد خود را ارسال کنید و در قرعه کشی برنده شوید.",
                            reply_markup=key_user_not_login)


@app.on_message(filters.private & filters.regex("^📝ثبت نام و ورود$"))
async def singIn_function(client, message):
    await message.reply("🔥به بخش ثبت نام و ورود خوش امدید.",
                        reply_markup=ReplyKeyboardMarkup([["بازگشت"]], resize_keyboard=True))
    if check_exist_id(file_user_login, message.from_user.id) == 1:
        await message.reply("✅شما قبلا ثبت نام کردید \n❕لطفا برای استفاده از ربات بر روی  /start کلیک کنید.❕",
                            reply_markup=key_user_not_login)
    else:
        name = await client.ask(message.from_user.id, "📝✏️نام و نام خانوادگی خود را وارد کنید:")
        if name.text == "بازگشت":
            await name.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_not_login)
        elif name.text in LIST_WORDS:
            await name.reply("❌نام و نام خانوادگی وارد شده اشتباه است. \n"
                             "⚠️لطفا دوباره بر روی '📝ثبت نام و ورود' کلیک کرده و مجدد ثبت نام کنید.",
                             reply_markup=key_user_not_login)
        else:
            number = await client.ask(name.from_user.id, "📞شماره تلفن خود را وارد کنید")
            if number.text == "بازگشت":
                await number.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_not_login)
            elif (number.text in LIST_WORDS) or len(number.text) <= 10:
                await number.reply("❌شماره تلفن وارد شده اشتباه است. \n"
                                   "لفطا دوباره بر روی '📝ثبت نام و ورود' کلیک و ثبت نام کنید.",
                                   reply_markup=key_user_not_login)
            else:
                try:
                    int(number.text)
                    save_id(file_user_login, number.from_user.id)
                    save_information(file_information, [number.from_user.id, name.text, number.text])
                    await number.reply("✅ثبت نام و ورود شما با موفقیت انجام شد.", reply_markup=key_user_login)
                except:
                    await number.reply(
                        "❌شماره تلفن وارد شده باید جزو اعداد باشد(0-9)\n"
                        "شماره وارد شده اشتباه است. لطفا دوباره بر روی"
                        " '📝ثبت نام و ورود' کلیک کرده و با دقت مراحل ثبت نام را تکمیل کنید.",
                        reply_markup=key_user_not_login)


@app.on_message(filters.private & filters.regex("^📝ثبت پیشنهادات و انتقادات$"))
async def criticism_function(client, message):
    if check_exist_id(file_user_login, message.from_user.id) == 1:
        information = get_information(message.from_user.id, file_information)
        await message.reply(f"""🤖سلام کاربر {information[1]} 
✳️به بخش پیشنهادات و انتقادات خوش آمدید🖐 : """,
                            reply_markup=ReplyKeyboardMarkup([["بازگشت"]], resize_keyboard=True))
        criticism = await client.ask(message.from_user.id, "▪️ لطفا پیشنهاد یا انتقاد خود را با ما در میان بگذارید⁉️")
        if criticism.text == "بازگشت":
            await criticism.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_login)
        elif criticism in LIST_WORDS:
            await criticism.reply("❌انتقاد یا پیشنهاد با فرمت اشتباهی وارد شده است\n"
                                  "لطفا دوباره با ارسال '📝ثبت پیشنهادات و انتقادات' انتقاد یاپیشنهاد خود را ثبت کنید.",
                                  reply_markup=key_user_login)
        else:
            date = jdatetime.date.today().strftime("%d-%m-%Y")
            await client.send_message(admin,
                                      f"**📝ثبت پیشنهادات و انتقادات**\n"
                                      f"📝نام و نام خانوادگی:{information[1]}\n📞تلفن:{information[2]}\n"
                                      f"📆تاریخ:{date}\n📝متن پیام:{criticism.text}", reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                "📩ذخیره", callback_data="save"
                            ),
                            InlineKeyboardButton(
                                "🏅برتر", callback_data="better"
                            )
                        ]
                    ]
                ))
            await criticism.reply("""✅انتقاد یا پیشنهاد شما ثبت شد.
🙏از اینکه به فکر دیار خود هستید متشکریم🙌
همچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی بهترین 
نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""", reply_markup=key_user_login)
    else:
        await message.reply("""🤖سلام کاربر ناشناس 
✳️به بخش پیشنهادات و انتقادات خوش آمدید🖐""",
                            reply_markup=ReplyKeyboardMarkup([["بازگشت"]], resize_keyboard=True))
        criticism = await client.ask(message.from_user.id, "▪️ لطفا پیشنهاد یا انتقاد خود را با ما در میان بگذارید⁉️")
        if criticism.text == "بازگشت":
            await criticism.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_login)
        elif criticism in LIST_WORDS:
            await criticism.reply("❌انتقاد یا پیشنهاد با فرمت اشتباهی وارد شده است\n"
                                  "لطفا دوباره با ارسال '📝ثبت پیشنهادات و انتقادات' انتقاد یاپیشنهاد خود را ثبت کنید.",
                                  reply_markup=key_user_login)
        else:
            date = jdatetime.date.today().strftime("%d-%m-%Y")
            await client.send_message(admin, f"**📝ثبت پیشنهادات و انتقادات**\n"
                                             f"از طرف کاربر 👤ناشناس\n📆تاریخ:{date}\ntext:{criticism.text}")
            await criticism.reply("""✅انتقاد یا پیشنهاد شما ثبت شد.
🙏از اینکه به فکر دیار خود هستید متشکریم🙌
همچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی بهترین 
نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""", reply_markup=key_user_login)


@app.on_message(filters.private & filters.regex("^🖋ثبت طرح📝$"))
async def proposal_function(client, message):
    if check_exist_id(file_user_login, message.from_user.id) == 1:
        information = get_information(message.from_user.id, file_information)
        await message.reply(f"""🤖سلام کاربر {information[1]} 
✳️به بخش ثبت طرح خوش آمدید🖐 : """, reply_markup=ReplyKeyboardMarkup(
            [
                ["بازگشت"]
            ], resize_keyboard=True
        ))
        proposal = await client.ask(message.from_user.id, "▪️ لطفا طرح خود را با ما در میان بگذارید⁉️")
        if proposal.text == "بازگشت":
            await proposal.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_login)
        elif proposal.text in LIST_WORDS:
            await proposal.reply("❌طرح با فرمت اشتباهی وارد شده است\n"
                                  "لطفا دوباره با ارسال '🖋ثبت طرح📝' طرح خود را ثبت کنید.",
                                 reply_markup=key_user_login)
        else:
            date = jdatetime.date.today().strftime("%d-%m-%Y")
            await client.send_message(admin,
                                      f"**📝ثبت طرح**\n"
                                      f"📝نام و نام خانوادگی:{information[1]}\n📞تلفن:{information[2]}\n"
                                      f"📆تاریخ:{date}\n📝متن پیام:{proposal.text}", reply_markup=InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                "📩ذخیره", callback_data="save"
                            ),
                            InlineKeyboardButton(
                                "🏅برتر", callback_data="better"
                            )
                        ]
                    ]
                ))
            await proposal.reply("""✅طرح شما ثبت شد.
🙏از اینکه به فکر دیار خود هستید متشکریم🙌
همچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی بهترین 
نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""", reply_markup=key_user_login)
    else:
        await message.reply("""🤖سلام کاربر ناشناس 
✳️به بخش ثبت طرح خوش آمدید🖐""", reply_markup=ReplyKeyboardMarkup(
            [
                ["بازگشت"]
            ], resize_keyboard=True
        ))
        proposal = await client.ask(message.from_user.id, "▪️ لطفا طرح خود را با ما در میان بگذارید⁉️")
        if proposal.text == "بازگشت":
            await proposal.reply("🔙به منو قبل بازگشتید", reply_markup=key_user_login)
        elif proposal.text in LIST_WORDS:
            await proposal.reply("❌طرح با فرمت اشتباهی وارد شده است\n"
                                  "لطفا دوباره با ارسال '🖋ثبت طرح📝' طرح خود را ثبت کنید.",
                                 reply_markup=key_user_login)
        else:
            date = jdatetime.date.today().strftime("%d-%m-%Y")
            await client.send_message(admin, f"**📝ثبت طرح**\n"
                                             f"از طرف کاربر 👤ناشناس\n📆تاریخ:{date}\ntext: {proposal.text}")
            await proposal.reply("""✅طرح شما ثبت شد.
🙏از اینکه به فکر دیار خود هستید متشکریم🙌
همچنین میتوانید با 📝ثبت نام و ورود و ارسال طرح و پیشنهادات خود که به پیشرفت روستا کمک کند در 🔶قرعه کشی بهترین 
نظرات و پیشنهادات شرکت کرده و به 💎قید قرعه برنده 🏆جوایز نفیس باشید.""", reply_markup=key_user_login)


@app.on_message(filters.private & filters.regex("^🤖درباره ربات$"))
async def about_bot(client, message):
    await message.reply(ABOUT_BOT)


@app.on_message(filters.private & filters.regex("^✅بازگشت به منو اولیه♻️$"))
async def back_to_main_meno(client, message):
    await message.reply("✅به منو اولیه برگشتید.♻️", reply_markup=key_user_not_login)


@app.on_message(filters.private & filters.regex("^📄گزارشکار پروژه های در حال اجرا ⛓$"))
async def work_project(client, message):
    await message.reply("""‼️ این بخش بدلیل توسعه و طراحی ربات در حال حاضر در دسترس نیست ❌

📣 در صورت در دسترس قرار گرفتن این بخش به شما اطلاع رسانی خواهد شد.""")
# run telegram bot
required()
app.run()
